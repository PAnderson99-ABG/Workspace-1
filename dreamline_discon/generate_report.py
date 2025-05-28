import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from datetime import datetime
import os

# === CONFIGURATION ===
headers = ['SKU', 'Organization', 'Parent SKU', 'Product Status', 'SKU Name*']
column_widths = [20, 25, 20, 20, 30]
header_fills = ['#00345B', '#00345B', '#00345B', '#C5D9F1', '#C5D9F1']
header_texts = ['#FFFFFF', '#FFFFFF', '#FFFFFF', '#00008B', '#00008B']

csv_input_path = "C:/Users/panderson/OneDrive - American Bath Group/Documents/Reports/CSV.DTP0799-Example.csv"

# === Generate filename with date (and save to OneDrive folder) ===
today_str = datetime.today().strftime("%Y-%m-%d")
output_path = f"C:/Users/panderson/OneDrive - American Bath Group/Documents/Reports/Discontinued SKUs - {today_str}.xlsx"

# === STEP 1: Read CSV (skip header row) ===
csv_data = pd.read_csv(csv_input_path, skiprows=1, header=None)
csv_data.columns = headers

# === STEP 2: Create Excel with header row ===
df = pd.DataFrame(columns=headers)
df.to_excel(output_path, index=False)

# === STEP 3: Append CSV data starting from row 2 ===
with pd.ExcelWriter(output_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
    csv_data.to_excel(writer, index=False, header=False, startrow=1)

# === STEP 4: Style headers ===
wb = load_workbook(output_path)
ws = wb.active

for col_num, (width, fill_hex, text_hex) in enumerate(zip(column_widths, header_fills, header_texts), start=1):
    cell = ws.cell(row=1, column=col_num)
    cell.fill = PatternFill(start_color=fill_hex.replace('#', ''), end_color=fill_hex.replace('#', ''), fill_type="solid")
    cell.font = Font(color=text_hex.replace('#', ''), bold=True)
    ws.column_dimensions[cell.column_letter].width = width

wb.save(output_path)
print(f"✅ Excel file created and saved to OneDrive for Power Automate: {output_path}")
