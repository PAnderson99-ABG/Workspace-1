import os
from dotenv import load_dotenv
import smartsheet
import pandas as pd
from datetime import datetime, timedelta
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# === LOAD API TOKEN FROM .env FILE ===
load_dotenv()
ACCESS_TOKEN = os.getenv('SMARTSHEET_API_TOKEN')

# === INITIALIZE SMARTSHEET CLIENT ===
smartsheet_client = smartsheet.Smartsheet(ACCESS_TOKEN)
smartsheet_client.errors_as_exceptions(True)

# === CONFIGURATION ===
SHEET_ID = int(os.getenv("PCM_REQUESTS_ID"))
DEALER_COLUMN_NAME = "Dealer"
EXPORT_FOLDER = r"C:/Users/panderson/OneDrive - American Bath Group/Documents/Reports"
ANDIE_EMAILS = {os.getenv("ANDIE_EMAIL").lower()}
BRIANNA_EMAILS = {os.getenv("BRIANNA_EMAIL").lower()}
WATCH_EMAILS = ANDIE_EMAILS.union(BRIANNA_EMAILS)
CUTOFF_DAYS = 14

# === FETCH SHEET ===
sheet = smartsheet_client.Sheets.get_sheet(SHEET_ID)

# === IDENTIFY COLUMN IDS ===
column_map = {col.title: col.id for col in sheet.columns}
dealer_col_id = column_map.get(DEALER_COLUMN_NAME)
date_requested_col_id = column_map.get("Date Requested")

if not dealer_col_id or not date_requested_col_id:
    raise Exception("Missing required column(s): 'Dealer' or 'Date Requested'")

# === FILTER ROWS WHERE 'Dealer' IS 'Home Depot' ===
matching_rows = []
for row in sheet.rows:
    for cell in row.cells:
        if cell.column_id == dealer_col_id and cell.value == "Home Depot":
            matching_rows.append(row)
            break

print(f"✅ Found {len(matching_rows)} rows for 'Home Depot'")

# === MAP COLUMN IDS TO TITLES ===
column_id_map = {col.id: col.title for col in sheet.columns}

# === BUILD DATA ===
data = []

for row in matching_rows:
    row_dict = {column_id_map.get(cell.column_id, f"Col_{cell.column_id}"): cell.value for cell in row.cells}

    # === FILTER BY STATUS ===
    status = str(row_dict.get("Status") or "").strip().lower()
    if status in ["complete", "cancelled", "submission error"]:
        print(f"⏭️ Row {row.id} skipped — status: {status}")
        continue

    # === FILTER BY REQUEST DATE ===
    request_date_value = row_dict.get("Date Requested")
    if request_date_value:
        try:
            if isinstance(request_date_value, str):
                try:
                    request_date = datetime.strptime(request_date_value.strip(), "%m/%d/%Y")
                except ValueError:
                    request_date = datetime.strptime(request_date_value.strip(), "%Y-%m-%d")
            elif isinstance(request_date_value, datetime):
                request_date = request_date_value
            elif hasattr(request_date_value, 'isoformat'):
                request_date = datetime.combine(request_date_value, datetime.min.time())
            else:
                print(f"⏭️ Row {row.id} skipped — unrecognized request date: {request_date_value}")
                continue

            if (datetime.now() - request_date).days < CUTOFF_DAYS:
                print(f"⏭️ Row {row.id} skipped — request date {request_date.date()} is too recent")
                continue
        except Exception as e:
            print(f"⏭️ Row {row.id} skipped — error parsing request date: {e}")
            continue
    else:
        print(f"⏭️ Row {row.id} skipped — missing 'Date Requested'")
        continue

    # === FETCH ALL COMMENTS WITH AUTHOR & TIMESTAMP ===
    all_comments = []
    latest_timestamp_raw = None
    latest_author = ""
    latest_request_to_watch = None
    latest_comment_from_watch = None

    try:
        discussions = smartsheet_client.Discussions.get_row_discussions(
            sheet_id=SHEET_ID,
            row_id=row.id,
            include="comments"
        ).data

        for discussion in discussions:
            for comment in discussion.comments:
                author = comment.created_by.name if comment.created_by else "Unknown"
                email = comment.created_by.email.lower() if comment.created_by and comment.created_by.email else ""
                comment_time = comment.created_at
                comment_text = comment.text.strip().replace('\n', ' ')

                # Collect formatted comment
                all_comments.append(f"{author} [{comment_time.strftime('%Y-%m-%d %H:%M:%S')}]: {comment_text}")

                # Track latest overall comment
                if not latest_timestamp_raw or comment_time > latest_timestamp_raw:
                    latest_timestamp_raw = comment_time
                    latest_author = author

                # Tag detection (if Andie or Brianna were mentioned)
                if any(watch_email in comment_text.lower() for watch_email in WATCH_EMAILS):
                    if not latest_request_to_watch or comment_time > latest_request_to_watch:
                        latest_request_to_watch = comment_time

                # Track replies from Andie or Brianna
                if email in WATCH_EMAILS:
                    if not latest_comment_from_watch or comment_time > latest_comment_from_watch:
                        latest_comment_from_watch = comment_time

        if not latest_timestamp_raw:
            print(f"⏭️ Row {row.id} skipped — no comments found")
            continue

        latest_timestamp = latest_timestamp_raw.strftime('%Y-%m-%d %H:%M:%S')
        all_comments_str = "; ".join(all_comments)

    except Exception as e:
        print(f"⏭️ Row {row.id} skipped — error fetching comments: {e}")
        continue

    # === Determine if row should be highlighted ===
    highlight_yellow = False
    if latest_request_to_watch:
        no_reply_or_old_reply = not latest_comment_from_watch or latest_comment_from_watch < latest_request_to_watch
        overdue = (datetime.now() - latest_request_to_watch.replace(tzinfo=None)).days >= CUTOFF_DAYS
        if no_reply_or_old_reply and overdue:
            highlight_yellow = True

    row_dict["All Comments"] = all_comments_str
    row_dict["Comment Author"] = latest_author
    row_dict["Comment Date"] = latest_timestamp
    row_dict["_highlight_yellow"] = highlight_yellow  # internal only, not exported

    print(f"✅ Row {row.id} added — comment date: {latest_timestamp}")
    data.append(row_dict)

# === EXPORT TO EXCEL ===
df = pd.DataFrame(data)
today_str = datetime.now().strftime("%Y-%m-%d")
filename = f"TEST_{today_str}.xlsx"
output_path = os.path.join(EXPORT_FOLDER, filename)

if not data:
    print("\n⚠️ No rows passed all filters — Excel file will be empty.")

# Export only visible columns
export_df = df.drop(columns=["_highlight_yellow"])
export_df.to_excel(output_path, index=False)

# === Apply Excel Highlighting ===
wb = load_workbook(output_path)
ws = wb.active
yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

highlight_flags = df["_highlight_yellow"].tolist()

for i, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=0):
    if highlight_flags[i]:
        for cell in row:
            cell.fill = yellow_fill

wb.save(output_path)
print(f"\n✅ Excel file saved with row highlights: {output_path}")
