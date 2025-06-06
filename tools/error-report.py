import re
import pandas as pd
from tkinter import Tk, filedialog
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook

# === File dialog to select .log file ===
Tk().withdraw()
file_path = filedialog.askopenfilename(
    title="Select the log file",
    filetypes=[("Log files", "*.log"), ("Text files", "*.txt"), ("All files", "*.*")]
)
if not file_path:
    raise Exception("No file selected.")

# === Read lines ===
with open(file_path, 'r', encoding='utf-8') as f:
    lines = [line.strip() for line in f if line.strip()]

rows = []

for line in lines:
    # Base match for Error Type and Product
    base_match = re.match(r"^(.*?): Product (\S+): (.*)", line)
    if base_match:
        error_type, product_number, remainder = base_match.groups()
        attempted_import = ""
        error_desc = ""

        # If '=' is present, split into error and attempted import
        if " = " in remainder:
            error_desc, attempted_import = remainder.split(" = ", 1)
            attempted_import = attempted_import.strip("' ")
        else:
            # Use the string after the last colon
            if remainder.count(":") >= 1:
                error_desc, attempted_import = remainder.rsplit(":", 1)
                attempted_import = attempted_import.strip()
            else:
                error_desc = remainder  # fallback

        rows.append({
            "Error Type": error_type.strip(),
            "Product Number": product_number.strip(),
            "Error Description": error_desc.strip(),
            "Attempted Import": attempted_import
        })

# === Write to Excel with autofit columns ===
if rows:
    df = pd.DataFrame(rows)
    output_path = file_path.replace(".log", "_parsed.xlsx").replace(".txt", "_parsed.xlsx")
    df.to_excel(output_path, index=False)

    wb = load_workbook(output_path)
    ws = wb.active
    for col in ws.columns:
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max_length + 2
    wb.save(output_path)

    print(f"✅ Excel file saved to:\n{output_path}")
else:
    print("⚠️ No matching error lines found.")
