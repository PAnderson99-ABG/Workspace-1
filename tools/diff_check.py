import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from tkinter import Tk, filedialog


def load_file_strict(filepath):
    ext = os.path.splitext(filepath)[1].lower()
    if ext == '.csv':
        return pd.read_csv(filepath, dtype=str, keep_default_na=False)
    elif ext in ['.xls', '.xlsx']:
        print(f"⚠️  Excel file detected: '{os.path.basename(filepath)}'. Only the first sheet will be compared.")
        return pd.read_excel(filepath, sheet_name=0, dtype=str).fillna('')
    else:
        raise ValueError(f"Unsupported file type: {ext}")


def pad_dataframe(df, length):
    missing = length - len(df)
    if missing > 0:
        empty_df = pd.DataFrame([[''] * df.shape[1]] * missing, columns=df.columns)
        df = pd.concat([df, empty_df], ignore_index=True)
    return df


def compare_rows(df1, df2):
    max_len = max(len(df1), len(df2))
    df1 = pad_dataframe(df1, max_len)
    df2 = pad_dataframe(df2, max_len)

    differing_rows_1 = []
    differing_rows_2 = []
    diffs = []

    for i in range(max_len):
        row1 = df1.iloc[i].values.tolist()
        row2 = df2.iloc[i].values.tolist()
        diff_cols = [j for j in range(max(len(row1), len(row2)))
                     if (j >= len(row1) or j >= len(row2)) or row1[j] != row2[j]]
        if diff_cols:
            differing_rows_1.append(row1)
            differing_rows_2.append(row2)
            diffs.append(diff_cols)

    return differing_rows_1, differing_rows_2, diffs


def write_to_excel_with_highlights(rows1, rows2, headers1, headers2, diffs, out_path):
    wb = Workbook()
    red_fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")

    def add_sheet(wb, title, data, headers, diffs_list):
        ws = wb.create_sheet(title)
        ws.append(list(headers))
        for r_idx, row in enumerate(data, start=2):
            for c_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                if r_idx - 2 < len(diffs_list) and c_idx - 1 in diffs_list[r_idx - 2]:
                    cell.fill = red_fill

    add_sheet(wb, "File1 Differences", rows1, headers1, diffs)
    add_sheet(wb, "File2 Differences", rows2, headers2, diffs)

    # Remove default sheet
    default_sheet = wb[wb.sheetnames[0]]
    wb.remove(default_sheet)

    wb.save(out_path)


def select_file(title):
    root = Tk()
    root.withdraw()
    file_path = filedialog.askopenfilename(title=title, filetypes=[
        ("CSV and Excel files", "*.csv *.xlsx *.xls"),
        ("CSV files", "*.csv"),
        ("Excel files", "*.xlsx *.xls")
    ])
    root.destroy()
    return file_path


def main():
    print("Please select the first file (CSV or Excel)")
    file1 = select_file("Select First File")
    print("Please select the second file (CSV or Excel)")
    file2 = select_file("Select Second File")

    if not file1 or not file2:
        print("File selection was cancelled.")
        return

    df1 = load_file_strict(file1)
    df2 = load_file_strict(file2)

    differing_rows_1, differing_rows_2, diffs = compare_rows(df1, df2)

    file1_name = os.path.splitext(os.path.basename(file1))[0]
    file2_name = os.path.splitext(os.path.basename(file2))[0]
    output_filename = f"diff_check.xlsx"

    # Hardcoded output path
    reports_folder = r"C:/Users/panderson/OneDrive - American Bath Group/Documents/Reports"
    os.makedirs(reports_folder, exist_ok=True)
    output_path = os.path.join(reports_folder, output_filename)

    write_to_excel_with_highlights(differing_rows_1, differing_rows_2, df1.columns, df2.columns, diffs, output_path)

    print(f"✅ Comparison complete. Output saved to '{output_path}'")


if __name__ == '__main__':
    main()
