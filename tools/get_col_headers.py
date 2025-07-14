import os
import csv
import tkinter as tk
from tkinter import messagebox
from dotenv import load_dotenv
import smartsheet
import pandas as pd
from openpyxl.styles import Font, PatternFill
from openpyxl import load_workbook

# === LOAD API TOKEN FROM .env ===
load_dotenv()
ACCESS_TOKEN = os.getenv("SMARTSHEET_API_TOKEN")

# === CONFIG ===
OPTIONS_FILE = "Master_Data_Sheets.csv"
OUTPUT_XLSX = "C:/Users/panderson/OneDrive - American Bath Group/Documents/Paul_Anderson/Reports/column_headers.xlsx"

# === LOAD SHEET OPTIONS ===
sheet_choices = []
try:
    with open(OPTIONS_FILE, mode='r', encoding='utf-8') as f:
        reader = csv.DictReader(f, fieldnames=["Name", "SheetID"])
        next(reader)  # Skip header
        sheet_choices = list(reader)
except FileNotFoundError:
    print(f"❌ Options file not found: {OPTIONS_FILE}")
    exit(1)

# === GUI FOR MULTI-SELECTION ===
def show_multi_select_gui():
    window = tk.Tk()
    window.title("Select Smartsheets")

    tk.Label(window, text="Select one or more sheets:", font=("Arial", 12, "bold")).pack(pady=10)

    checkbox_vars = []
    for i, row in enumerate(sheet_choices):
        var = tk.BooleanVar()
        cb = tk.Checkbutton(window, text=row["Name"], variable=var)
        cb.pack(anchor="w")
        checkbox_vars.append(var)

    def on_continue():
        if not any(var.get() for var in checkbox_vars):
            messagebox.showerror("Selection Error", "Please select at least one sheet.")
        else:
            window.quit()
            window.destroy()

    tk.Button(window, text="Continue", command=on_continue).pack(pady=10)
    window.mainloop()

    return [i for i, var in enumerate(checkbox_vars) if var.get()]

selected_indexes = show_multi_select_gui()

if not selected_indexes:
    print("❌ No sheets selected. Exiting.")
    exit(1)

# === INITIALIZE SMARTSHEET CLIENT ===
smartsheet_client = smartsheet.Smartsheet(ACCESS_TOKEN)
smartsheet_client.errors_as_exceptions(True)

# === FETCH HEADERS FOR SELECTED SHEETS ===
excel_data = {}
for idx in selected_indexes:
    sheet_name = sheet_choices[idx]["Name"]
    sheet_id = int(sheet_choices[idx]["SheetID"])

    try:
        sheet = smartsheet_client.Sheets.get_sheet(sheet_id)
        headers = [col.title for col in sheet.columns if not col.hidden]
        excel_data[sheet_name[:31]] = headers  # Excel tab names max 31 chars
        print(f"✅ Fetched headers for: {sheet_name}")
    except smartsheet.exceptions.ApiError as e:
        print(f"❌ API Error for '{sheet_name}': {e.error.result_code} - {e.error.message}")

# === EXPORT TO XLSX FILE ===
if excel_data:
    # Write the headers to Excel (no formatting yet)
    with pd.ExcelWriter(OUTPUT_XLSX, engine="openpyxl") as writer:
        for sheet_name, headers in excel_data.items():
            df = pd.DataFrame([headers])
            df.to_excel(writer, sheet_name=sheet_name, index=False, header=False)

    # Reopen file for styling
    wb = load_workbook(OUTPUT_XLSX)
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    white_bold_font = Font(color="FFFFFF", bold=True)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row[:9]:  # First 9 columns (A–I)
                cell.fill = red_fill
                cell.font = white_bold_font

    wb.save(OUTPUT_XLSX)

    print(f"\n✅ Headers exported and styled in: {OUTPUT_XLSX}")
else:
    print("⚠️ No data exported due to errors.")
