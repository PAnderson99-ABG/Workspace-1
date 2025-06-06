import os
import pandas as pd
import tkinter as tk
from tkinter import filedialog, simpledialog, messagebox
from openpyxl import load_workbook

# === GUI setup ===
root = tk.Tk()
root.withdraw()

# === File picker ===
file_path = filedialog.askopenfilename(title="Select Excel file", filetypes=[("Excel files", "*.xlsx")])
if not file_path:
    raise Exception("No file selected.")

# === Read all sheet names ===
xl = pd.ExcelFile(file_path)
sheets = xl.sheet_names

# === Sheet selector dialog ===
selected_sheets = simpledialog.askstring(
    "Sheet Selection",
    f"Available sheets:\n{', '.join(sheets)}\n\nEnter comma-separated sheet names to process:"
)
if not selected_sheets:
    raise Exception("No sheets selected.")
selected_sheets = [s.strip() for s in selected_sheets.split(",") if s.strip() in sheets]

# === Output path ===
output_path = filedialog.asksaveasfilename(title="Save As", defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
if not output_path:
    raise Exception("No output path provided.")

# === Load original workbook ===
wb = load_workbook(file_path)

# === Process each selected sheet ===
for sheet_name in selected_sheets:
    if sheet_name not in wb.sheetnames:
        messagebox.showwarning("Sheet Skipped", f"Skipping '{sheet_name}' — not found in workbook.")
        continue

    df = xl.parse(sheet_name)

    if 'salsify:data_inheritance_hierarchy_level_id' not in df.columns or 'Unique ID' not in df.columns:
        messagebox.showwarning("Sheet Skipped", f"Skipping '{sheet_name}' — missing required columns.")
        continue

    is_base = df['salsify:data_inheritance_hierarchy_level_id'] == 'Base'
    is_sellable = df['salsify:data_inheritance_hierarchy_level_id'] == 'Sellable'
    columns_to_copy = df.columns[10:]

    base_df = df[is_base].copy()
    sellable_df = df[is_sellable].copy()
    sellable_groups = sellable_df.groupby('Parent ID')

    # Map Unique ID to row number (Excel is 1-based + header row)
    uid_to_rownum = {
        row['Unique ID']: idx + 2
        for idx, row in df.iterrows()
        if row['salsify:data_inheritance_hierarchy_level_id'] == 'Base'
    }

    ws = wb[sheet_name]
    header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    col_start = 11  # Column K = 11 in 1-based Excel

    for parent_id, group in sellable_groups:
        if parent_id not in uid_to_rownum:
            continue

        first = group.iloc[0]
        row_num = uid_to_rownum[parent_id]
        row_cells = list(ws.iter_rows(min_row=row_num, max_row=row_num))[0]

        for i, col_name in enumerate(columns_to_copy):
            if col_start + i - 1 < len(row_cells):
                cell = row_cells[col_start + i - 1]
                cell.value = first.get(col_name)

# === Save and open updated workbook ===
wb.save(output_path)
os.startfile(output_path)
messagebox.showinfo("Done", f"File saved and opened:\n{output_path}")
